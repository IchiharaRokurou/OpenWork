# -*- coding: utf-8 -*-

import sys
import time
import chromedriver_binary
from selenium import webdriver
from selenium.webdriver.common.by import By
import re 
from openpyxl import Workbook
from openpyxl import load_workbook

"""
import 
pip install openpyxl
pip install selenium
pip install chromedriver-binary==104.0.5112.79
"""

# ログインページのURL
url = "https://www.vorkers.com/login.php"
top_url = "https://www.vorkers.com/my_top"
base_url = "https://www.vorkers.com"

# Webドライバ
driver = webdriver.Chrome()
driver.get(url) 

# ユーザ情報 入力してください
USERNAME = "fordevelop6960@gmail.com"
PASSWORD = "vFKfwrr2M@ipEL"


def main():
    # URLを叩き、htmlを表示    
    driver.get(url) 
    time.sleep(0.5)
    login()
    search(["株式会社学研"])
    get_search_results()
    
    

# ログイン
def login():
    # メールとパスワードを入力
    mail = driver.find_element("name","_username") 
    password = driver.find_element("name","_password") 
    
    # クリア
    mail.clear()
    password.clear()
    # フォーム入力
    mail.send_keys(USERNAME)
    password.send_keys(PASSWORD)
    # エンター
    mail.submit()



def search(name:str=""):
    """
    オープンワークで企業名検索する関数
    Args:
        name (str, optional): 検索したい名前. Defaults to "".
    """
    # トップページに戻ってから
    back_to_mypage()
    search_box = driver.find_element("name","src_str")
    # 検索    
    search_box.send_keys(name)
    search_box.submit()


# 検索結果取得
def get_search_results():
    results = driver.find_elements(By.CLASS_NAME,"searchCompanyName")
    pattern = r'<div>(.*?)</div>'
    a_tags = []
    # 各要素から名前のaダグのみ取得
    for result in results:
        content = result.get_attribute('innerHTML').replace('\n', '').replace('\r', '').replace('\t', '').replace(' ', '')
        text = re.match(pattern, content)
        # aタグからurlと名前取得
        text = re.match(r'<div>.*<ahref="(.*?)">(.*?)</a>.*?', text.group())
        if text:
            a_tags.append(text.groups())
    # 各ページ移動
    for a_tag in a_tags:
        comp_url, name = a_tag
        print(name)
        driver.get(base_url + comp_url)
        save_info(name, base_url + comp_url)
        time.sleep(5)

# xlsxファイル作成&書き込み
def make_xlsx(filename, scores:dict, review_num, all_review_dict:dict):
    # ワークブックの新規作成と保存
    wb = Workbook()
    wb.save('{}.xlsx'.format(filename))
    wb = load_workbook('{}.xlsx'.format(filename))
    ws = wb.active
    # ワークシートの追記
    ws.title = "Summary"
    
    # スコア
    ws.cell(row=1, column=1).value = '項目'
    ws.cell(row=1, column=2).value = 'スコア'
    for i, key in enumerate(scores.keys()):
        ws.cell(row=i+2, column=1).value = key
        ws.cell(row=i+2, column=2).value = scores[key]
    # 口コミ数
    ws.cell(row=1, column=3).value = '口コミ数'
    ws.cell(row=2, column=3).value = review_num
    wb.save('{}.xlsx'.format(filename))
    
    # 次に口コミについて記入
    for i, key in enumerate(all_review_dict.keys()):
        wb.create_sheet(key)
        wb.active = i + 1
        ws = wb.active
        ws.cell(row=1, column=1).value = '番号'
        ws.cell(row=1, column=2).value = '日付'
        ws.cell(row=1, column=3).value = 'プロフィール'
        ws.cell(row=1, column=4).value = 'スコア'
        ws.cell(row=1, column=5).value = '口コミ'
        each_review_dict = all_review_dict[key]
        for j in range(1, max([int(x) for x in each_review_dict.keys()]) + 1):
            personal_dict = each_review_dict[j]
            ws.cell(row=1+j, column=1).value = j
            ws.cell(row=1+j, column=2).value = personal_dict['date']
            ws.cell(row=1+j, column=3).value = personal_dict["profile"]
            ws.cell(row=1+j, column=4).value = personal_dict["score"]
            ws.cell(row=1+j, column=5).value = personal_dict["article_answer"]
    wb.save('{}.xlsx'.format(filename))



# 情報保存
def save_info(comp_name, comp_url):
    # スコア辞書取得
    scores = get_scores()
    review_num = get_review_nums()
    all_review_dict = get_all_review(comp_url.replace('company', 'company_answer'))
    make_xlsx(comp_name, scores, review_num, all_review_dict)
    
    
def get_scores():
    scores = driver.find_element(By.CLASS_NAME,'scoreList-8')
    # <li> ~ </li>に分解
    contents = scores.get_attribute('innerHTML').replace('\n', '').replace('\r', '').replace('\t', '').replace(' ', '').split('</li>')
    score_dict = dict()
    # 各liのデータ取得
    pattern = r'<li.*?><dl>(.*?)</dl>'
    for content in contents:
        result = re.match(pattern, content)
        # スコアと点数を取得
        if result:
            table_name, score = re.match(r'.*?<dt.*?>(.*?)</dt><dd.*?>(.*?)</dd>.*?', result.group()).groups()
            print(table_name, score)
            score_dict[table_name] = score
    return score_dict


# 口コミ数取得
def get_review_nums():
    review = driver.find_element(By.XPATH, '//h2[contains(text(),"カテゴリ")]')
    content = review.get_attribute('innerHTML').replace('\n', '').replace('\r', '').replace('\t', '').replace(' ', '')
    pattern = r'.*<span>(.*?)</span>.*'
    result = re.match(pattern, content)
    return result.group(1)


# 口コミへ
def get_all_review(review_url):
    driver.get(review_url)
    all_review_dict = dict()
    for no in range(1, 10 + 1):
        if no == 7:
            continue
        review_name, review_dict = get_each_review_data(review_url + '&q_no={}'.format(no))
        all_review_dict[review_name] = review_dict
    time.sleep(1)
    return all_review_dict


# 各口コミ
def get_each_review_data(review_url):
    driver.get(review_url)
    review_num = driver.find_element(By.XPATH, "//a[@href='{}']".format(review_url.replace(base_url, '')))
    content = review_num.get_attribute('innerHTML').replace('\n', '').replace('\r', '').replace('\t', '').replace(' ', '')
    pattern = r'<span.*?>(.*?)<span.*?>（(.*?)件）</span></span>'
    result = re.match(pattern, content)
    # レビューの数とカテゴリ取得
    review_dict = dict()
    if not result:
        pattern = r'(.*?)<span.*?>.*?>(.*?)件</span>.*?</span>'
        result = re.match(pattern, content)
    if result:
        review_name, num = result.groups()
        review_name = review_name.replace('［', '').replace('］', '')
        i = 1
        # 次ページへも遷移
        num = int(num)
        if num == 0:
            return review_name, {}

        # 25ずつ表示
        max_page = 2 + num//25 if num % 25 == 0 else 2 + 1 + num//25
        for page in range(2, max_page):
            #　口コミ取得
            reviews = driver.find_elements(By.CLASS_NAME,'article')
            for review in reviews:
                # 日付取得
                time_tag = review.find_element(By.TAG_NAME, 'time')
                datetime = time_tag.get_attribute("datetime")
                profile = review.find_element(By.CSS_SELECTOR, ".mr-5.v-m").text
                score = review.find_element(By.CSS_SELECTOR, ".ml-5.fs-14").text
                content = review.find_element(By.CLASS_NAME, 'article_answer').get_attribute('innerHTML').replace('\n', '').replace('\r', '').replace('\t', '').replace(' ', '')
                article_answer = ''
                if '<span' not in content:
                    article_answer = content
                else:
                    pattern = r'(.*?)<span.*?>(.*?)</span>'
                    result = re.match(pattern, content)
                    if result:
                        article_answer = ''.join(result.groups())
                # print(datetime)
                # print(profile)
                # print(score)
                # print(article_answer)
                review_dict[i] = {"date": datetime, "profile": profile, "score": score, "article_answer": article_answer.replace('<br>', '')}
                i += 1
            
            # 次ページへ
            driver.get(review_url + '&next_page={}'.format(page))
            time.sleep(1)
        return review_name, review_dict
    else:
        return "Error", {}


# マイページ戻る
def back_to_mypage():
    driver.get(top_url) 
    time.sleep(0.5)


if __name__ == '__main__':
    main()